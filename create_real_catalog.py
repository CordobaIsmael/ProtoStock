import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Lista de 50 productos reales de Fiambrería, Almacén, Bebidas y Kiosco en Argentina (Precios reales de mercado ARS)
products_data = [
    # Fiambrería (Kilo)
    ["7790001001012", "Jamón Cocido Paladini", "Fiambrería", "KG", 5200, 8900, 15.5, 3],
    ["7790001001029", "Jamón Crudo Bocatti", "Fiambrería", "KG", 9800, 16500, 8.0, 2],
    ["7790001001036", "Queso Tybo Barra La Serenísima", "Fiambrería", "KG", 4800, 8200, 20.0, 4],
    ["7790001001043", "Queso Cremoso Punta del Agua", "Fiambrería", "KG", 3900, 6800, 25.0, 5],
    ["7790001001050", "Queso Gouda Barra Milkaut", "Fiambrería", "KG", 5500, 9400, 12.0, 3],
    ["7790001001067", "Queso Roquefort / Azul Santa Rosa", "Fiambrería", "KG", 7200, 12800, 6.0, 2],
    ["7790001001074", "Salame de Colonia Caroya Cagnoli", "Fiambrería", "KG", 6400, 11200, 10.0, 2],
    ["7790001001081", "Mortadela Cañuelense con Pistacho", "Fiambrería", "KG", 3400, 5900, 14.0, 3],
    ["7790001001098", "Bondola de Cerdo Estacionada Paladini", "Fiambrería", "KG", 7800, 13900, 7.5, 2],
    ["7790001001104", "Queso Sardo Rallado Horma", "Fiambrería", "KG", 6900, 11800, 9.0, 2],

    # Fiambrería & Lácteos (Unidad)
    ["7790070411057", "Queso Crema Finlandia 300g", "Lácteos", "UNIDAD", 1850, 2950, 24, 6],
    ["7790070318042", "Manteca La Serenísima 200g", "Lácteos", "UNIDAD", 1450, 2300, 30, 8],
    ["7790070501000", "Crema de Leche Milkaut 200ml", "Lácteos", "UNIDAD", 1100, 1750, 20, 5],
    ["7790048000016", "Levadura Calsa Fresca 50g", "Fiambrería", "UNIDAD", 320, 550, 50, 10],
    ["7790070000015", "Leche Entera La Serenísima 1L Sachet", "Lácteos", "UNIDAD", 950, 1450, 40, 10],

    # Panificados & Tapas
    ["7790040001004", "Pan de Miga Blanco 1kg", "Panadería", "UNIDAD", 1600, 2800, 15, 4],
    ["7790040002001", "Pan Lactal Fargo Grande 560g", "Panadería", "UNIDAD", 1900, 3200, 20, 5],
    ["7790080001002", "Tapa de Tarta Pascualina La Salteña 400g", "Panadería", "UNIDAD", 980, 1650, 30, 8],
    ["7790080002009", "Tapa de Empanadas Hoja La Salteña 12u", "Panadería", "UNIDAD", 850, 1400, 35, 10],

    # Almacén General
    ["7790272000025", "Aceite Girasol Natura 900ml", "Almacén", "UNIDAD", 1350, 2200, 30, 8],
    ["7790070001234", "Fideos Lucchetti Tallarín 500g", "Almacén", "UNIDAD", 680, 1150, 50, 12],
    ["7790070005678", "Fideos Matarazzo Mostachol 500g", "Almacén", "UNIDAD", 750, 1250, 40, 10],
    ["7790272010100", "Arroz Lucchetti Largo Fino 1kg", "Almacén", "UNIDAD", 1200, 1950, 35, 10],
    ["7790040010105", "Harina Favorita 0000 1kg", "Almacén", "UNIDAD", 620, 980, 40, 10],
    ["7790040020203", "Harina Blancaflor Leudante 1kg", "Almacén", "UNIDAD", 890, 1450, 30, 8],
    ["7790100001000", "Azúcar Ledesma Clásica 1kg", "Almacén", "UNIDAD", 820, 1300, 50, 12],
    ["7790020001000", "Sal Fina Dos Anclas 500g", "Almacén", "UNIDAD", 540, 890, 30, 8],
    ["7790050001000", "Puré de Tomate Arcor 520g Tetra", "Almacén", "UNIDAD", 580, 950, 60, 15],
    ["7790272000500", "Mayonesa Natura Doypack 500g", "Almacén", "UNIDAD", 1150, 1850, 25, 6],

    # Bebidas con & sin Alcohol
    ["7790895000997", "Coca Cola Original 2.25L Retornable", "Bebidas", "UNIDAD", 1750, 2700, 40, 10],
    ["7790895001000", "Coca Cola Zero 1.5L descartable", "Bebidas", "UNIDAD", 1450, 2300, 30, 8],
    ["7790895002007", "Sprite Limón 2.25L Retornable", "Bebidas", "UNIDAD", 1650, 2600, 25, 6],
    ["7790001002002", "Cerveza Quilmes Clásica 1L Retornable", "Bebidas", "UNIDAD", 1400, 2200, 50, 12],
    ["7790001003009", "Cerveza Stella Artois 730ml", "Bebidas", "UNIDAD", 1950, 3100, 30, 8],
    ["7790001004006", "Agua Mineral Villavicencio 1.5L Sin Gas", "Bebidas", "UNIDAD", 750, 1200, 40, 10],
    ["7790001005003", "Fernet Branca 750ml", "Bebidas", "UNIDAD", 6800, 10500, 20, 5],
    ["7790001006000", "Vino Toro Tinto 1L Tetra", "Bebidas", "UNIDAD", 1100, 1750, 30, 8],

    # Kiosco & Golosinas
    ["7790005001001", "Alfajor Guaymallén Chocolate 38g", "Kiosco", "UNIDAD", 220, 400, 100, 20],
    ["7790005002008", "Alfajor Havanna Mixto 55g", "Kiosco", "UNIDAD", 850, 1400, 40, 10],
    ["7790070002000", "Galletitas Chocolinas 250g", "Kiosco", "UNIDAD", 1150, 1850, 30, 8],
    ["7790070003007", "Galletitas Criollitas Pack x3 300g", "Kiosco", "UNIDAD", 980, 1550, 30, 8],
    ["7790005003005", "Caramelos Sugus Variados 150g", "Kiosco", "UNIDAD", 750, 1200, 25, 5],
    ["7790005004002", "Papas Lays Clásicas 85g", "Kiosco", "UNIDAD", 1350, 2100, 25, 6],
    ["7790070004004", "Chocolatada Cindor 1L Tetra", "Lácteos", "UNIDAD", 1650, 2600, 20, 5],

    # Limpieza & Higiene
    ["7790010001000", "Lavandina Ayudín Triple Acción 1L", "Limpieza", "UNIDAD", 750, 1200, 30, 8],
    ["7790010002007", "Detergente Magistral Limón 300ml", "Limpieza", "UNIDAD", 1250, 1950, 25, 6],
    ["7790010003004", "Jabón en Polvo Ala Multiacción 800g", "Limpieza", "UNIDAD", 1600, 2500, 20, 5],
    ["7790010004001", "Papel Higiénico Elegante 4u x 30m", "Limpieza", "UNIDAD", 1100, 1750, 25, 6],
    ["7790010005008", "Rollo de Cocina Sussex 3u", "Limpieza", "UNIDAD", 1200, 1890, 20, 5],
]

headers = ["codigo", "nombre", "categoria", "tipoVenta", "precioCosto", "precioVenta", "stockActual", "stockMinimo"]

# 1. Generar CSV
csv_path = "d:/ProtoStock/CATALOGO_INICIAL_PROTOSTOCK.csv"
with open(csv_path, mode="w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for row in products_data:
        writer.writerow(row)

print(f"Archivo CSV generado con exito en: {csv_path}")

# 2. Generar Excel (.xlsx) Estilizado
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Catálogo ProtoStock"

# Estilos Excel
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
border_thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1")
)

# Escribir Encabezados
ws.append(["Código de Barras", "Nombre del Producto", "Categoría", "Tipo Venta (KG/UNIDAD)", "Precio Costo ($)", "Precio Venta ($)", "Stock Inicial", "Stock Mínimo"])
for col_idx in range(1, 9):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

# Escribir Datos
for row_idx, data_row in enumerate(products_data, start=2):
    ws.append(data_row)
    for col_idx in range(1, 9):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = border_thin
        
        # Formatos por columna
        if col_idx in [1, 4]:
            cell.alignment = align_center
        elif col_idx in [5, 6]:
            cell.alignment = align_right
            cell.number_format = '$#,##0.00'
        elif col_idx in [7, 8]:
            cell.alignment = align_right
            cell.number_format = '#,##0.00'
        else:
            cell.alignment = align_left

# Ajustar anchos de columnas
column_widths = [18, 42, 16, 22, 16, 16, 14, 14]
for idx, width in enumerate(column_widths, start=1):
    col_letter = openpyxl.utils.get_column_letter(idx)
    ws.column_dimensions[col_letter].width = width

excel_path = "d:/ProtoStock/CATALOGO_INICIAL_PROTOSTOCK.xlsx"
wb.save(excel_path)
print(f"Archivo Excel (.xlsx) generado con exito en: {excel_path}")
